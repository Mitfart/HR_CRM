import csv
import io
import json
import urllib.parse
import urllib.request
import uuid
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
import openpyxl
from pydantic import BaseModel
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from google.auth.transport.requests import Request
from google.oauth2.service_account import Credentials

from app.config import settings
from app.database import get_db
from app.dependencies import require_manager
from app.models.application import Application
from app.models.app_setting import AppSetting
from app.models.user import User
from app.tasks_sheets import sync_google_sheets

router = APIRouter()
DEALS_COMMENTS_KEY = "deals_cell_comments"
DEALS_OVERRIDES_KEY = "deals_cell_overrides"
DEALS_KANBAN_KEY_PREFIX = "deals_kanban_v1"
DEALS_ARCHIVE_KEY_PREFIX = "deals_archive_v1"

DEFAULT_DEAL_STAGES = [
    ("Новая заявка", "#2563eb"),
    ("Формирование запроса", "#7c3aed"),
    ("Оплачен аванс", "#059669"),
    ("Подбор анкет", "#0891b2"),
    ("Заключение договора", "#ea580c"),
    ("Отморозились", "#64748b"),
    ("Собеседования", "#db2777"),
    ("Пробные дни", "#9333ea"),
    ("Основной счет на доплату", "#ca8a04"),
    ("Доплата от соискателя", "#16a34a"),
    ("Финальный счёт", "#0f766e"),
    ("Запрос отзыва о работе с компанией", "#475569"),
]


class DealCommentIn(BaseModel):
    gid: str
    row: int
    col: int
    text: str


class DealCellIn(BaseModel):
    gid: str
    row: int
    col: int
    value: str


class DealKanbanStageIn(BaseModel):
    gid: str = "1396224654"
    name: str
    color: str | None = None


class DealKanbanStageUpdateIn(BaseModel):
    gid: str = "1396224654"
    name: str | None = None
    color: str | None = None
    sort_order: int | None = None


class DealKanbanItemUpdateIn(BaseModel):
    gid: str = "1396224654"
    stage_id: str


class DealArchiveRestoreIn(BaseModel):
    gid: str = "1396224654"
    item_id: str
    restored: bool = True


def _kanban_key(gid: str) -> str:
    return f"{DEALS_KANBAN_KEY_PREFIX}:{gid}"


def _archive_key(gid: str) -> str:
    return f"{DEALS_ARCHIVE_KEY_PREFIX}:{gid}"


def _default_kanban_state() -> dict:
    return {
        "stages": [
            {
                "id": f"stage_{idx + 1}",
                "name": name,
                "color": color,
                "sort_order": idx,
                "is_won": name in {"Финальный счёт", "Запрос отзыва о работе с компанией"},
                "is_lost": name == "Отморозились",
            }
            for idx, (name, color) in enumerate(DEFAULT_DEAL_STAGES)
        ],
        "assignments": {},
        "manual_items": [],
    }


def _normalize_kanban_state(value: dict | None) -> dict:
    state = value if isinstance(value, dict) else {}
    default_state = _default_kanban_state()
    stages = state.get("stages") if isinstance(state.get("stages"), list) else default_state["stages"]
    assignments = state.get("assignments") if isinstance(state.get("assignments"), dict) else {}
    manual_items = state.get("manual_items") if isinstance(state.get("manual_items"), list) else []

    normalized_stages = []
    for idx, stage in enumerate(stages):
        if not isinstance(stage, dict):
            continue
        stage_id = str(stage.get("id") or f"stage_{idx + 1}")
        name = str(stage.get("name") or "").strip()
        if not name:
            continue
        normalized_stages.append(
            {
                "id": stage_id,
                "name": name,
                "color": stage.get("color") or "#6366f1",
                "sort_order": int(stage.get("sort_order", idx)),
                "is_won": bool(stage.get("is_won", False)),
                "is_lost": bool(stage.get("is_lost", False)),
            }
        )

    if not normalized_stages:
        normalized_stages = default_state["stages"]

    normalized_stages = sorted(normalized_stages, key=lambda item: item["sort_order"])
    for idx, stage in enumerate(normalized_stages):
        stage["sort_order"] = idx

    return {
        "stages": normalized_stages,
        "assignments": {str(k): str(v) for k, v in assignments.items()},
        "manual_items": manual_items,
    }


def _normalize_archive_state(value: dict | None) -> dict:
    state = value if isinstance(value, dict) else {}
    auto = state.get("auto") if isinstance(state.get("auto"), list) else []
    restored = state.get("restored") if isinstance(state.get("restored"), list) else []
    return {
        "auto": sorted({str(item_id) for item_id in auto if item_id}),
        "restored": sorted({str(item_id) for item_id in restored if item_id}),
    }


async def _get_kanban_state(db: AsyncSession, gid: str) -> dict:
    row = await db.get(AppSetting, _kanban_key(gid))
    state = _normalize_kanban_state(row.value if row else None)
    if not row:
        db.add(AppSetting(key=_kanban_key(gid), value=state))
        await db.flush()
    elif row.value != state:
        row.value = state
        await db.flush()
    return state


async def _save_kanban_state(db: AsyncSession, gid: str, state: dict) -> dict:
    normalized = _normalize_kanban_state(state)
    row = await db.get(AppSetting, _kanban_key(gid))
    if row:
        row.value = normalized
    else:
        db.add(AppSetting(key=_kanban_key(gid), value=normalized))
    await db.flush()
    return normalized


async def _get_archive_state(db: AsyncSession, gid: str) -> dict:
    row = await db.get(AppSetting, _archive_key(gid))
    state = _normalize_archive_state(row.value if row else None)
    if not row:
        db.add(AppSetting(key=_archive_key(gid), value=state))
        await db.flush()
    elif row.value != state:
        row.value = state
        await db.flush()
    return state


async def _save_archive_state(db: AsyncSession, gid: str, state: dict) -> dict:
    normalized = _normalize_archive_state(state)
    row = await db.get(AppSetting, _archive_key(gid))
    if row:
        row.value = normalized
    else:
        db.add(AppSetting(key=_archive_key(gid), value=normalized))
    await db.flush()
    return normalized


def _row_item_title(row: list[str], source_idx: int) -> str:
    candidates = [row[idx].strip() for idx in (2, 1, 3, 0) if idx < len(row) and row[idx].strip()]
    return candidates[0] if candidates else f"Сделка #{source_idx + 1}"


def _row_item_amount(row: list[str]) -> int | None:
    for value in row:
        digits = "".join(ch for ch in str(value) if ch.isdigit())
        if len(digits) >= 4:
            try:
                return int(digits)
            except ValueError:
                return None
    return None


def _build_kanban_items(rows: list[list[str]], state: dict, gid: str) -> list[dict]:
    stages = state["stages"]
    default_stage_id = stages[0]["id"]
    assignments = state["assignments"]
    items = []
    for source_idx, row in enumerate(rows[1:], start=1):
        if not any(str(cell).strip() for cell in row):
            continue
        item_id = f"sheet:{gid}:{source_idx}"
        items.append(
            {
                "id": item_id,
                "title": _row_item_title(row, source_idx),
                "stage_id": assignments.get(item_id, default_stage_id),
                "amount": _row_item_amount(row),
                "owner_id": None,
                "created_at": datetime.utcnow().isoformat(),
                "source": "google_sheet",
                "source_row": source_idx,
                "client": row[2].strip() if len(row) > 2 else "",
                "description": row[3].strip() if len(row) > 3 else "",
            }
        )
    items.extend(item for item in state["manual_items"] if isinstance(item, dict))
    return items


def _is_gray_hex(color: str | None) -> bool:
    if not color:
        return False
    hex_color = color.strip().lower().lstrip("#")
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    if len(hex_color) != 6:
        return False
    try:
        red = int(hex_color[0:2], 16)
        green = int(hex_color[2:4], 16)
        blue = int(hex_color[4:6], 16)
    except ValueError:
        return False
    spread = max(red, green, blue) - min(red, green, blue)
    brightness = (red + green + blue) / 3
    return spread <= 22 and 70 <= brightness <= 245


def _archived_row_ids_from_styles(styles: dict[str, dict[str, str]], gid: str) -> set[str]:
    archived_rows: set[int] = set()
    for key, style in styles.items():
        if not isinstance(style, dict):
            continue
        row_text = str(key).split(":", 1)[0]
        if not row_text.isdigit():
            continue
        row_idx = int(row_text)
        if row_idx <= 0:
            continue
        if _is_gray_hex(style.get("bg")):
            archived_rows.add(row_idx)
    return {f"sheet:{gid}:{row_idx}" for row_idx in archived_rows}


def _fetch_sheet_rows(gid: str) -> list[list[str]]:
    spreadsheet_id = "1f1Uldq5ncCNRGmrRjtyFf88-CgV2_DzVVmh_bf-Z73U"
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        raw = resp.read().decode("utf-8")
    return list(csv.reader(io.StringIO(raw)))


def _resolve_sheet_title(spreadsheet_id: str, gid: str, token: str) -> str | None:
    meta_url = (
        f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}"
        f"?fields=sheets(properties(sheetId,title))"
    )
    meta_req = urllib.request.Request(
        meta_url,
        headers={"Authorization": f"Bearer {token}", "User-Agent": "Mozilla/5.0"},
    )
    with urllib.request.urlopen(meta_req, timeout=30) as meta_resp:
        meta = json.loads(meta_resp.read().decode("utf-8"))
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {})
        if str(props.get("sheetId")) == gid:
            return props.get("title")
    return None


def _color_to_hex(color_obj: dict | None) -> str | None:
    if not color_obj:
        return None
    r = int(round(float(color_obj.get("red", 1)) * 255))
    g = int(round(float(color_obj.get("green", 1)) * 255))
    b = int(round(float(color_obj.get("blue", 1)) * 255))
    return f"#{r:02x}{g:02x}{b:02x}"


def _fetch_sheet_notes_and_styles(spreadsheet_id: str, gid: str) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    if not settings.google_credentials_file:
        return {}, {}
    try:
        creds = Credentials.from_service_account_file(
            settings.google_credentials_file,
            scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
        )
        creds.refresh(Request())
        sheet_title = _resolve_sheet_title(spreadsheet_id, gid, creds.token)
        if not sheet_title:
            return {}, {}

        range_name = urllib.parse.quote(sheet_title)
        grid_url = (
            f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}"
            f"?includeGridData=true&ranges={range_name}"
            f"&fields=sheets(data(rowData(values(note,effectiveFormat(backgroundColor,textFormat.foregroundColor)))))"
        )
        grid_req = urllib.request.Request(
            grid_url,
            headers={"Authorization": f"Bearer {creds.token}", "User-Agent": "Mozilla/5.0"},
        )
        with urllib.request.urlopen(grid_req, timeout=30) as notes_resp:
            notes_data = json.loads(notes_resp.read().decode("utf-8"))

        notes_map: dict[str, str] = {}
        styles_map: dict[str, dict[str, str]] = {}
        for sheet in notes_data.get("sheets", []):
            for data_block in sheet.get("data", []):
                for r_idx, row_data in enumerate(data_block.get("rowData", [])):
                    for c_idx, cell in enumerate(row_data.get("values", [])):
                        key = f"{r_idx}:{c_idx}"
                        note = (cell.get("note") or "").strip()
                        if note:
                            notes_map[key] = note

                        fmt = cell.get("effectiveFormat") or {}
                        bg_hex = _color_to_hex(fmt.get("backgroundColor"))
                        txt_hex = _color_to_hex((fmt.get("textFormat") or {}).get("foregroundColor"))
                        style: dict[str, str] = {}
                        if bg_hex and bg_hex.lower() != "#ffffff":
                            style["bg"] = bg_hex
                        if txt_hex and txt_hex.lower() != "#000000":
                            style["fg"] = txt_hex
                        if style:
                            styles_map[key] = style
        return notes_map, styles_map
    except Exception:
        return _fetch_sheet_styles_from_xlsx(spreadsheet_id, gid)


def _argb_to_hex(argb: str | None) -> str | None:
    if not argb:
        return None
    code = str(argb).strip()
    if len(code) == 8:
        code = code[2:]
    if len(code) != 6:
        return None
    return f"#{code.lower()}"


def _sheet_title_from_gid(gid: str) -> str:
    # Known mapping for this workbook.
    if gid == "1396224654":
        return "вакансии 26"
    return "вакансии 26"


def _fetch_sheet_styles_from_xlsx(spreadsheet_id: str, gid: str) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    """
    Fallback style extraction without Google API credentials.
    Reads workbook styles from XLSX export.
    """
    try:
        xlsx_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
        req = urllib.request.Request(xlsx_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            xlsx_bytes = resp.read()

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        title = _sheet_title_from_gid(gid)
        ws = wb[title] if title in wb.sheetnames else wb[wb.sheetnames[0]]

        styles_map: dict[str, dict[str, str]] = {}
        for row_idx, row in enumerate(ws.iter_rows()):
            for col_idx, cell in enumerate(row):
                style: dict[str, str] = {}

                fill = cell.fill
                if fill and fill.fill_type == "solid":
                    bg_hex = _argb_to_hex(getattr(fill.fgColor, "rgb", None))
                    if bg_hex and bg_hex != "#ffffff":
                        style["bg"] = bg_hex

                font = cell.font
                if font:
                    fg_hex = _argb_to_hex(getattr(font.color, "rgb", None))
                    if fg_hex and fg_hex != "#000000":
                        style["fg"] = fg_hex

                if style:
                    styles_map[f"{row_idx}:{col_idx}"] = style

        return {}, styles_map
    except Exception:
        return {}, {}


@router.get("")
async def list_deals(
    skip: int = 0,
    limit: int = 200,
    search: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    query = (
        select(Application)
        .where(
            Application.manager_notes.ilike("[deal-sync]%")
        )
        .order_by(Application.updated_at.desc())
        .offset(skip)
        .limit(limit)
    )
    if search:
        s = f"%{search.strip()}%"
        query = query.where(
            or_(
                Application.description.ilike(s),
                Application.max_contact.ilike(s),
            )
        )
    result = await db.execute(query)
    return result.scalars().all()


@router.post("/sync-now")
async def sync_deals_now(_: User = Depends(require_manager)):
    task = sync_google_sheets.delay()
    return {"status": "queued", "task_id": task.id}


@router.get("/sheet-snapshot")
async def get_sheet_snapshot(
    gid: str = "1396224654",
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    spreadsheet_id = "1f1Uldq5ncCNRGmrRjtyFf88-CgV2_DzVVmh_bf-Z73U"
    rows = _fetch_sheet_rows(gid)
    notes, styles = _fetch_sheet_notes_and_styles(spreadsheet_id, gid)
    comments_row = await db.get(AppSetting, DEALS_COMMENTS_KEY)
    local_comments = comments_row.value if comments_row and isinstance(comments_row.value, dict) else {}
    overrides_row = await db.get(AppSetting, DEALS_OVERRIDES_KEY)
    local_overrides = overrides_row.value if overrides_row and isinstance(overrides_row.value, dict) else {}
    archive_state = await _get_archive_state(db, gid)
    archive_state["auto"] = sorted(_archived_row_ids_from_styles(styles, gid))
    archive_state = await _save_archive_state(db, gid, archive_state)
    return {
        "gid": gid,
        "rows": rows,
        "notes": notes,
        "styles": styles,
        "local_comments": local_comments,
        "local_overrides": local_overrides,
        "archive": {
            "auto_item_ids": archive_state["auto"],
            "restored_item_ids": archive_state["restored"],
        },
    }


@router.get("/kanban")
async def get_deals_kanban(
    gid: str = "1396224654",
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    state = await _get_kanban_state(db, gid)
    rows = _fetch_sheet_rows(gid)
    archive_state = await _get_archive_state(db, gid)
    auto_archived_ids = set(archive_state["auto"])
    restored_ids = set(archive_state["restored"])
    items = [
        item
        for item in _build_kanban_items(rows, state, gid)
        if item.get("id") not in auto_archived_ids or item.get("id") in restored_ids
    ]
    return {
        "gid": gid,
        "stages": state["stages"],
        "items": items,
    }


@router.post("/kanban/stages")
async def create_deals_kanban_stage(
    payload: DealKanbanStageIn,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Stage name is required")
    state = await _get_kanban_state(db, payload.gid)
    stage = {
        "id": f"stage_{uuid.uuid4().hex[:10]}",
        "name": name,
        "color": payload.color or "#6366f1",
        "sort_order": len(state["stages"]),
        "is_won": False,
        "is_lost": False,
    }
    state["stages"].append(stage)
    state = await _save_kanban_state(db, payload.gid, state)
    return {"ok": True, "stage": stage, "stages": state["stages"]}


@router.patch("/kanban/stages/{stage_id}")
async def update_deals_kanban_stage(
    stage_id: str,
    payload: DealKanbanStageUpdateIn,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    state = await _get_kanban_state(db, payload.gid)
    stage = next((item for item in state["stages"] if item["id"] == stage_id), None)
    if not stage:
        raise HTTPException(status_code=404, detail="Stage not found")
    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Stage name is required")
        stage["name"] = name
    if payload.color is not None:
        stage["color"] = payload.color or "#6366f1"
    if payload.sort_order is not None:
        target_order = max(0, min(payload.sort_order, len(state["stages"]) - 1))
        reordered = [item for item in state["stages"] if item["id"] != stage_id]
        reordered.insert(target_order, stage)
        state["stages"] = reordered
    state = await _save_kanban_state(db, payload.gid, state)
    return {"ok": True, "stage": next(item for item in state["stages"] if item["id"] == stage_id), "stages": state["stages"]}


@router.delete("/kanban/stages/{stage_id}")
async def delete_deals_kanban_stage(
    stage_id: str,
    gid: str = "1396224654",
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    state = await _get_kanban_state(db, gid)
    if len(state["stages"]) <= 1:
        raise HTTPException(status_code=400, detail="At least one stage must remain")
    stage_ids = {stage["id"] for stage in state["stages"]}
    if stage_id not in stage_ids:
        raise HTTPException(status_code=404, detail="Stage not found")
    fallback_stage_id = next(stage["id"] for stage in state["stages"] if stage["id"] != stage_id)
    state["stages"] = [stage for stage in state["stages"] if stage["id"] != stage_id]
    state["assignments"] = {
        item_id: (fallback_stage_id if assigned_stage_id == stage_id else assigned_stage_id)
        for item_id, assigned_stage_id in state["assignments"].items()
    }
    for item in state["manual_items"]:
        if isinstance(item, dict) and item.get("stage_id") == stage_id:
            item["stage_id"] = fallback_stage_id
    state = await _save_kanban_state(db, gid, state)
    return {"ok": True, "stages": state["stages"]}


@router.patch("/kanban/items/{item_id}")
async def update_deals_kanban_item(
    item_id: str,
    payload: DealKanbanItemUpdateIn,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    state = await _get_kanban_state(db, payload.gid)
    stage_ids = {stage["id"] for stage in state["stages"]}
    if payload.stage_id not in stage_ids:
        raise HTTPException(status_code=404, detail="Stage not found")
    state["assignments"][item_id] = payload.stage_id
    state = await _save_kanban_state(db, payload.gid, state)
    return {"ok": True, "item_id": item_id, "stage_id": payload.stage_id}


@router.post("/sheet-archive/restore")
async def restore_deal_from_archive(
    payload: DealArchiveRestoreIn,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    state = await _get_archive_state(db, payload.gid)
    restored = set(state["restored"])
    if payload.restored:
        restored.add(payload.item_id)
    else:
        restored.discard(payload.item_id)
    state["restored"] = sorted(restored)
    state = await _save_archive_state(db, payload.gid, state)
    return {"ok": True, "item_id": payload.item_id, "restored": payload.item_id in state["restored"]}


@router.post("/sheet-comments")
async def add_sheet_comment(
    payload: DealCommentIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_manager),
):
    key = f"{payload.gid}:{payload.row}:{payload.col}"
    row = await db.get(AppSetting, DEALS_COMMENTS_KEY)
    # Copy JSON payload to ensure SQLAlchemy detects changes.
    store = dict(row.value) if row and isinstance(row.value, dict) else {}
    comments_for_cell = list(store.get(key, []))
    if not isinstance(comments_for_cell, list):
        comments_for_cell = []
    comments_for_cell.append(
        {
            "text": payload.text.strip(),
            "author": user.full_name or user.email,
            "created_at": datetime.utcnow().isoformat(),
        }
    )
    store[key] = comments_for_cell
    if row:
        row.value = store
    else:
        db.add(AppSetting(key=DEALS_COMMENTS_KEY, value=store))
    await db.flush()
    return {"ok": True, "key": key, "count": len(comments_for_cell)}


@router.post("/sheet-cells")
async def upsert_sheet_cell(
    payload: DealCellIn,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_manager),
):
    key = f"{payload.gid}:{payload.row}:{payload.col}"
    row = await db.get(AppSetting, DEALS_OVERRIDES_KEY)
    # Copy JSON payload to ensure SQLAlchemy detects changes.
    store = dict(row.value) if row and isinstance(row.value, dict) else {}
    store[key] = payload.value
    if row:
        row.value = store
    else:
        db.add(AppSetting(key=DEALS_OVERRIDES_KEY, value=store))
    await db.flush()
    return {"ok": True, "key": key, "value": payload.value}
