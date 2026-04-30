"""
One-time import script: loads all data from the Google Sheets XLSX export into the CRM.

Usage (from backend/ directory):
    python scripts/import_from_sheets.py --xlsx /path/to/file.xlsx

The script:
  1. Creates the "Вакансии" pipeline + stages (if not existing)
  2. Imports вакансии 24/25/26 + ранжирование  → Deals
  3. Imports ⭐️НУР, 🎀НАСТЯ, 🧸АНЯ ХАБ, 💎ЛЕРА, ОТЧ Аня №1, актуализация базы → Candidates
  4. Skips ПАРОЛИ sheet entirely
  5. Records every row in sheets_sync_rows for future delta-sync
"""

import argparse
import hashlib
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# Make sure the app package is importable
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parents[2] / ".env")

from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.config import settings
from app.models.candidate import Candidate
from app.models.crm_deal import Deal
from app.models.crm_pipeline import Pipeline, Stage
from app.models.sheets_sync import SheetSyncRow

# ── Sheet classification (same as tasks_sheets.py) ───────────────────────────

VACANCY_SHEETS = {"вакансии 26", "вакансии 25", "вакансии 24", "ранжирование"}
CANDIDATE_SHEETS = {
    "⭐️НУР⭐️", "🎀 НАСТЯ🎀", "🧸 АНЯ ХАБ🧸", "💎ЛЕРА💎",
    "ОТЧ Аня №1", "актуализация базы",
}
SKIP_SHEETS = {"ПАРОЛИ", "%"}

EMOJI_STATUS = {
    "✅✔️": "Закрыт",
    "✅‼️": "Срочно",
    "✔️":   "Закрыт",
    "‼️":   "Срочно",
    "✅":   "В работе",
}

SOURCE_MAP = {
    "тг": "Telegram", "хх": "HeadHunter",
    "пм": "Помогатор", "pm": "Помогатор", "hh": "HeadHunter",
}

PIPELINE_STAGES = [
    ("Новая",    "#a0aec0", False, False),
    ("В работе", "#4299e1", False, False),
    ("Срочно",   "#ed8936", False, False),
    ("Закрыт",   "#48bb78", True,  False),
    ("Отменён",  "#fc8181", False, True),
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def row_hash(row: list[Any]) -> str:
    text = "|".join(str(c) if c is not None else "" for c in row)
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def strip_emojis(text: str) -> str:
    return re.sub(r"^[\u2705\u2714\ufe0f\u203c\ufe0f\u2b50\ufe0f\s]+", "", text).strip()


def detect_status(raw: str) -> str:
    for emoji, stage in EMOJI_STATUS.items():
        if raw.startswith(emoji):
            return stage
    if "✅" in raw and "✔️" in raw:
        return "Закрыт"
    if "‼️" in raw:
        return "Срочно"
    if "✅" in raw:
        return "В работе"
    return "Новая"


def parse_sources(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[/\s,]+", str(raw).lower())
    return [SOURCE_MAP[p] for p in parts if p in SOURCE_MAP]


def parse_salary(raw: Any) -> float | None:
    if raw is None:
        return None
    s = str(raw).replace("\u00a0", "").replace(" ", "").replace(".", "")
    nums = re.findall(r"\d+", s)
    return float(nums[0]) if nums else None


def normalize_phone(raw: Any) -> str | None:
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) == 11 and digits[0] == "8":
        digits = "7" + digits[1:]
    return "+" + digits if len(digits) >= 10 else None


def to_date(val: Any) -> datetime | None:
    if isinstance(val, datetime):
        return val
    if val is None:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%-d.%-m.%Y", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            pass
    return None


# ── Pipeline setup ────────────────────────────────────────────────────────────

def ensure_pipeline(session: Session) -> tuple[Pipeline, dict[str, Stage]]:
    pipeline = session.execute(
        select(Pipeline).where(Pipeline.name == "Вакансии", Pipeline.entity_type == "deal")
    ).scalar_one_or_none()

    if not pipeline:
        pipeline = Pipeline(name="Вакансии", entity_type="deal", is_default=True)
        session.add(pipeline)
        session.flush()
        print("  Created pipeline: Вакансии")

    stages_list = session.execute(
        select(Stage).where(Stage.pipeline_id == pipeline.id)
    ).scalars().all()
    stages = {s.name: s for s in stages_list}

    for order, (name, color, is_won, is_lost) in enumerate(PIPELINE_STAGES):
        if name not in stages:
            s = Stage(
                pipeline_id=pipeline.id,
                name=name,
                sort_order=order,
                color=color,
                is_won=is_won,
                is_lost=is_lost,
            )
            session.add(s)
            session.flush()
            stages[name] = s
            print(f"  Created stage: {name}")

    return pipeline, stages


# ── Vacancy import ────────────────────────────────────────────────────────────

def import_vacancy_row(
    session: Session,
    row: list[Any],
    row_idx: int,
    sheet_name: str,
    spreadsheet_id: str,
    pipeline: Pipeline,
    stages: dict[str, Stage],
    owner_cache: dict[str, Any],
    stats: dict,
):
    col = list(row) + [None] * 10
    date_raw, manager_raw, client_raw, source_raw, schedule_raw, pos_raw, salary_raw, loc_raw = col[:8]

    if not client_raw:
        return
    client_str = str(client_raw).strip()
    if not client_str or client_str.lower() in ("клиент", "client", "кому"):
        return
    # Skip month/year header rows like "Январь", "Февраль" etc.
    if re.match(r"^[А-Яа-яЁё]+$", client_str) and len(client_str) < 15:
        return

    h = row_hash(row)

    # Check if already tracked
    sync_rec = session.execute(
        select(SheetSyncRow).where(
            SheetSyncRow.spreadsheet_id == spreadsheet_id,
            SheetSyncRow.sheet_name == sheet_name,
            SheetSyncRow.row_number == row_idx,
        )
    ).scalar_one_or_none()

    if sync_rec and sync_rec.row_hash == h and sync_rec.entity_id:
        return  # unchanged

    status = detect_status(client_str)
    client_name = strip_emojis(client_str)
    sources = parse_sources(str(source_raw) if source_raw else None)
    salary = parse_salary(salary_raw)
    order_date = to_date(date_raw)

    # Fix typos in year (e.g. 16.11.2026 should be 2025, 06.01.0226 → 2026)
    if order_date and order_date.year > datetime.now().year + 1:
        fixed_year = datetime.now().year
        order_date = order_date.replace(year=fixed_year)

    manager_name = str(manager_raw).strip() if manager_raw else None
    stage = stages.get(status, stages.get("Новая"))

    # Look up owner
    owner_id = None
    if manager_name:
        key = manager_name.lower()
        if key in owner_cache:
            owner_id = owner_cache[key]

    title = f"{client_name} — {str(pos_raw).strip() if pos_raw else 'вакансия'}".strip(" —")

    custom = {
        "sheet_name": sheet_name,
        "spreadsheet_id": spreadsheet_id,
        "order_date": order_date.isoformat() if order_date else None,
        "manager_name": manager_name,
        "client_raw": client_str,
        "source_channels": sources,
        "schedule": str(schedule_raw).strip() if schedule_raw else None,
        "position": str(pos_raw).strip() if pos_raw else None,
        "salary_text": str(salary_raw).strip() if salary_raw else None,
        "location": str(loc_raw).strip() if loc_raw else None,
        "order_status": status,
    }

    if sync_rec and sync_rec.entity_id:
        deal = session.get(Deal, sync_rec.entity_id)
        if deal:
            deal.title = title
            deal.amount = salary
            deal.pipeline_id = pipeline.id
            deal.stage_id = stage.id if stage else None
            deal.owner_id = owner_id
            deal.custom_fields = custom
            sync_rec.row_hash = h
            sync_rec.last_synced_at = datetime.now(timezone.utc)
            stats["deals_updated"] += 1
            return

    # Duplicate guard: same client_raw + sheet
    existing = session.execute(
        select(Deal).where(
            Deal.custom_fields["client_raw"].astext == client_str,
            Deal.custom_fields["sheet_name"].astext == sheet_name,
        )
    ).scalar_one_or_none()

    if existing:
        deal = existing
    else:
        deal = Deal(
            title=title,
            amount=salary,
            pipeline_id=pipeline.id,
            stage_id=stage.id if stage else None,
            owner_id=owner_id,
            custom_fields=custom,
        )
        session.add(deal)
        session.flush()
        stats["deals_created"] += 1

    if sync_rec:
        sync_rec.entity_id = deal.id
        sync_rec.row_hash = h
        sync_rec.last_synced_at = datetime.now(timezone.utc)
    else:
        session.add(SheetSyncRow(
            spreadsheet_id=spreadsheet_id,
            sheet_name=sheet_name,
            row_number=row_idx,
            row_hash=h,
            entity_type="deal",
            entity_id=deal.id,
        ))


# ── Candidate import ──────────────────────────────────────────────────────────

def import_candidate_row(
    session: Session,
    row: list[Any],
    row_idx: int,
    sheet_name: str,
    spreadsheet_id: str,
    stats: dict,
):
    col = list(row) + [None] * 10

    # актуализация базы has: num, name, age, position, phone, source, notes
    # Manager sheets have:   num, date, name, age, spec, phone, site, result
    # Detect by checking if col[1] looks like a date or a name
    if sheet_name == "актуализация базы":
        _num, name_raw, age_raw, spec_raw, phone_raw, site_raw, result_raw = col[0], col[1], col[2], col[3], col[4], col[5], col[6]
    else:
        _num, _date, name_raw, age_raw, spec_raw, phone_raw, site_raw, result_raw = col[:8]

    if not name_raw:
        return
    name_str = str(name_raw).strip()
    if not name_str or name_str.lower() in ("фио", "имя", "name", "ф.и.о"):
        return
    # Skip rows that are section headers (e.g. "МАРТ", "АПРЕЛЬ")
    if re.match(r"^[А-ЯA-Z\s]+$", name_str) and len(name_str) < 20 and not any(c.islower() for c in name_str):
        return

    h = row_hash(row)

    sync_rec = session.execute(
        select(SheetSyncRow).where(
            SheetSyncRow.spreadsheet_id == spreadsheet_id,
            SheetSyncRow.sheet_name == sheet_name,
            SheetSyncRow.row_number == row_idx,
        )
    ).scalar_one_or_none()

    if sync_rec and sync_rec.row_hash == h and sync_rec.entity_id:
        return

    phone = normalize_phone(phone_raw)
    spec = str(spec_raw).strip() if spec_raw else None
    site = str(site_raw).strip() if site_raw else None
    result = str(result_raw).strip() if result_raw else None

    age: int | None = None
    if age_raw:
        try:
            age = int(float(str(age_raw)))
        except (ValueError, TypeError):
            pass

    contacts: dict = {}
    if phone:
        contacts["phone"] = phone

    notes_parts = []
    if site:
        notes_parts.append(f"Источник: {site}")
    if result:
        notes_parts.append(f"Результат: {result}")
    notes_parts.append(f"Лист: {sheet_name}")
    notes_text = "\n".join(notes_parts)

    tags = [t for t in [spec, sheet_name] if t]

    if sync_rec and sync_rec.entity_id:
        cand = session.get(Candidate, sync_rec.entity_id)
        if cand:
            cand.full_name = name_str
            cand.age = age
            cand.specialization = spec
            cand.contacts = contacts
            cand.notes = notes_text
            cand.tags = tags
            sync_rec.row_hash = h
            sync_rec.last_synced_at = datetime.now(timezone.utc)
            stats["candidates_updated"] += 1
            return

    # Duplicate guard by phone
    existing = None
    if phone:
        existing = session.execute(
            select(Candidate).where(Candidate.contacts["phone"].astext == phone)
        ).scalar_one_or_none()

    if existing:
        cand = existing
    else:
        cand = Candidate(
            full_name=name_str,
            age=age,
            specialization=spec,
            contacts=contacts,
            notes=notes_text,
            tags=tags,
        )
        session.add(cand)
        session.flush()
        stats["candidates_created"] += 1

    if sync_rec:
        sync_rec.entity_id = cand.id
        sync_rec.row_hash = h
        sync_rec.last_synced_at = datetime.now(timezone.utc)
    else:
        session.add(SheetSyncRow(
            spreadsheet_id=spreadsheet_id,
            sheet_name=sheet_name,
            row_number=row_idx,
            row_hash=h,
            entity_type="candidate",
            entity_id=cand.id,
        ))


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import Google Sheets XLSX into CRM")
    parser.add_argument("--xlsx", required=True, help="Path to exported .xlsx file")
    parser.add_argument(
        "--spreadsheet-id",
        default="1f1Uldq5ncCNRGmrRjtyFf88-CgV2_DzVVmh_bf-Z73U",
        help="Google Spreadsheet ID (used as foreign key in sync table)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}")
        sys.exit(1)

    spreadsheet_id = args.spreadsheet_id
    db_url = settings.database_url.replace("+asyncpg", "")

    print(f"Connecting to DB: {db_url[:40]}...")
    engine = create_engine(db_url, pool_pre_ping=True)

    print(f"Opening XLSX: {xlsx_path}")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    stats = {
        "deals_created": 0, "deals_updated": 0,
        "candidates_created": 0, "candidates_updated": 0,
        "skipped": 0, "errors": 0,
    }

    with Session(engine) as session:
        # 1. Ensure pipeline + stages exist
        print("\n[1/3] Setting up pipeline...")
        pipeline, stages = ensure_pipeline(session)
        session.commit()

        # Build owner cache (login → id)
        from app.models.user import User
        users = session.execute(select(User)).scalars().all()
        owner_cache = {}
        for u in users:
            if u.full_name:
                owner_cache[u.full_name.lower().strip()] = u.id
            if u.email:
                owner_cache[u.email.lower().strip()] = u.id

        # 2. Import vacancies → Deals
        print("\n[2/3] Importing vacancies (Deals)...")
        for sheet_name in wb.sheetnames:
            if sheet_name not in VACANCY_SHEETS:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            print(f"  Sheet '{sheet_name}': {len(rows)} rows")
            for row_idx, row in enumerate(rows):
                if row_idx == 0:
                    continue  # header
                if not any(c for c in row if c is not None):
                    continue  # empty
                try:
                    import_vacancy_row(
                        session, list(row), row_idx, sheet_name,
                        spreadsheet_id, pipeline, stages, owner_cache, stats,
                    )
                except Exception as exc:
                    stats["errors"] += 1
                    print(f"    ERROR row {row_idx}: {exc}")
            session.commit()
            print(f"    Deals created: {stats['deals_created']}, updated: {stats['deals_updated']}")

        # 3. Import candidates
        print("\n[3/3] Importing candidates...")
        for sheet_name in wb.sheetnames:
            if sheet_name not in CANDIDATE_SHEETS:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            print(f"  Sheet '{sheet_name}': {len(rows)} rows")
            for row_idx, row in enumerate(rows):
                if row_idx == 0:
                    continue
                if not any(c for c in row if c is not None):
                    continue
                try:
                    import_candidate_row(
                        session, list(row), row_idx, sheet_name,
                        spreadsheet_id, stats,
                    )
                except Exception as exc:
                    stats["errors"] += 1
                    print(f"    ERROR row {row_idx}: {exc}")
            session.commit()
            print(f"    Candidates created: {stats['candidates_created']}, updated: {stats['candidates_updated']}")

    engine.dispose()

    print("\n" + "=" * 50)
    print("IMPORT COMPLETE")
    print(f"  Deals created:       {stats['deals_created']}")
    print(f"  Deals updated:       {stats['deals_updated']}")
    print(f"  Candidates created:  {stats['candidates_created']}")
    print(f"  Candidates updated:  {stats['candidates_updated']}")
    print(f"  Errors:              {stats['errors']}")
    print("=" * 50)


if __name__ == "__main__":
    main()
